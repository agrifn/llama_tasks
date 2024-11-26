# generate_reports

import os
import datetime
from db_connection import get_db_connection
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import logging

# Configure logging
logging.basicConfig(
    filename='/path/to/your/logfile.log',
    level=logging.INFO,
    format='%(asctime)s %(levelname)s:%(message)s'
)

# Load environment variables
load_dotenv('/path/to/your/.env')  # Update the path to your .env file

def fetch_people():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        query = "SELECT person_id, name FROM people ORDER BY name;"
        cursor.execute(query)
        people = cursor.fetchall()
        connection.close()
        return people
    except Exception as e:
        logging.error("Error fetching people data", exc_info=True)
        return []

def fetch_tasks():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        query = "SELECT task_id, task_name FROM tasks ORDER BY task_name;"
        cursor.execute(query)
        tasks = cursor.fetchall()
        connection.close()
        return tasks
    except Exception as e:
        logging.error("Error fetching tasks data", exc_info=True)
        return []

def fetch_task_completions():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        query = """
        SELECT
            tc.person_id,
            tc.task_id,
            tc.completion_date
        FROM
            task_completion tc;
        """
        cursor.execute(query)
        completions = cursor.fetchall()
        connection.close()
        return completions
    except Exception as e:
        logging.error("Error fetching task completions data", exc_info=True)
        return []

def generate_excel_report(people, tasks, completions, report_file):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Task Completion Report"

        # Bold font for headers
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Write task names in the first row starting from column B
        for col_num, task in enumerate(tasks, start=2):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = task['task_name']
            ws[f"{col_letter}1"].font = bold_font
            ws[f"{col_letter}1"].alignment = center_alignment

        # Write people names in the first column starting from row 2
        for row_num, person in enumerate(people, start=2):
            ws[f"A{row_num}"] = person['name']
            ws[f"A{row_num}"].font = bold_font
            ws[f"A{row_num}"].alignment = Alignment(horizontal='left', vertical='center')

        # Create a dictionary for quick lookup of completion dates
        completion_dict = {}
        for completion in completions:
            key = (completion['person_id'], completion['task_id'])
            completion_dict[key] = completion['completion_date']

        # Fill in the completion dates
        person_id_to_row = {person['person_id']: row_num for row_num, person in enumerate(people, start=2)}
        task_id_to_col = {task['task_id']: col_num for col_num, task in enumerate(tasks, start=2)}

        date_format = 'YYYY-MM-DD'

        for (person_id, task_id), completion_date in completion_dict.items():
            row_num = person_id_to_row.get(person_id)
            col_num = task_id_to_col.get(task_id)
            if row_num and col_num:
                col_letter = get_column_letter(col_num)
                cell = ws[f"{col_letter}{row_num}"]
                cell.value = completion_date.strftime('%Y-%m-%d')
                cell.alignment = center_alignment

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook
        wb.save(report_file)
        logging.info(f"Report generated: {report_file}")
        print(f"Report generated: {report_file}")
    except Exception as e:
        logging.error("Error generating Excel report", exc_info=True)
        print("An error occurred while generating the report.")

def send_report(report_file):
    smtp_server = os.getenv('SMTP_SERVER')
    smtp_port = int(os.getenv('SMTP_PORT'))
    email_address = os.getenv('EMAIL_ADDRESS')
    email_password = os.getenv('EMAIL_PASSWORD')
    admin_emails = os.getenv('ADMIN_EMAILS')  # Comma-separated list of admin emails

    if not all([smtp_server, smtp_port, email_address, email_password, admin_emails]):
        logging.error("SMTP or admin email configuration is incomplete.")
        print("SMTP or admin email configuration is incomplete in the .env file.")
        return

    admin_email_list = [email.strip() for email in admin_emails.split(',')]

    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(email_address, email_password)
    except Exception as e:
        logging.error("Failed to connect to the SMTP server", exc_info=True)
        print(f"Failed to connect to the SMTP server: {e}")
        return

    msg = MIMEMultipart()
    msg['From'] = email_address
    msg['To'] = ', '.join(admin_email_list)
    msg['Subject'] = 'Task Completion Report'

    body = 'Please find the attached task completion report.'

    msg.attach(MIMEText(body, 'plain'))

    # Attach the report file
    attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    try:
        with open(report_file, 'rb') as file:
            attachment.set_payload(file.read())
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', f'attachment; filename={report_file}')
        msg.attach(attachment)
    except Exception as e:
        logging.error("Failed to read the report file", exc_info=True)
        print("Failed to read the report file.")
        return

    try:
        server.send_message(msg)
        logging.info(f"Report emailed to administrators: {admin_email_list}")
        print(f"Report emailed to administrators: {admin_email_list}")
    except Exception as e:
        logging.error("Failed to send report email", exc_info=True)
        print(f"Failed to send report email: {e}")
    finally:
        server.quit()

if __name__ == '__main__':
    try:
        # Fetch data
        people = fetch_people()
        tasks = fetch_tasks()
        completions = fetch_task_completions()

        if not people or not tasks:
            logging.error("No people or tasks data to generate report.")
            print("No data available to generate the report.")
        else:
            # Generate report
            report_file = 'task_report.xlsx'
            generate_excel_report(people, tasks, completions, report_file)

            # Send the report via email
            send_report(report_file)
    except Exception as e:
        logging.error("An error occurred in the main execution block", exc_info=True)
        print("An error occurred while running the report generation.")
